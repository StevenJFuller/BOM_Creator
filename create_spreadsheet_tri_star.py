from os import listdir
import tabula
import csv
from PyPDF2 import PdfFileReader, PdfFileWriter
from openpyxl import Workbook, load_workbook
from fractions import Fraction

def seperate_drawings(): # seperates drawings so that all drawings are single page only. Adds sheet number.
    print("Separating multi-page drawings...")
    drawing_names = listdir('drawings')
    for i in range(len(drawing_names)):
        inputpdf = PdfFileReader(open('drawings/' + drawing_names[i], 'rb'))
        for j in range(inputpdf.numPages):
            output = PdfFileWriter()
            output.addPage(inputpdf.getPage(j))
            with open('drawings_separated/' + drawing_names[i][:-4] + "_sheet_" + str(j + 1) + '.pdf', 'wb') as outputStream:
                output.write(outputStream)
    print("Done.")

def create_spreadhseet(): # sets up MTO spreadsheet with drawing names (typically ISO numbers) along the top row.
    print("Creating spreadsheet template...")
    pdf_files = []
    sep_dwg_names = listdir('drawings')
    for i in range(len(sep_dwg_names)):
        if sep_dwg_names[i][-4:] == '.pdf':
            pdf_files.append(sep_dwg_names[i][:-4])
    wb = load_workbook('Blank MTO r1.xlsx')
    ws = wb.active
    current_column = 23
    for i in range(len(pdf_files)):
        ws.cell(row=5, column=current_column, value=pdf_files[i])
        current_column += 1

    wb.save('MTO.xlsx')
    print("Done.")

def convert_to_csv(): # converts seperated drawings' BOM's to csv format --note: 'area' may need to be manipulated using tabula-java to define BOM list on drawing
    print("Converting PDFs into readable format...")
    tabula.convert_into_by_batch("drawings", output_format="csv", stream = True, area=[22.185,933.3,582.165,1187.28], guess=False)
    print("Done.")

def read_length(data): # Needed if lengths are in `ft'-in"` format
    
    if '-' in data:
        split_data = data.split("-")
        ft = float(split_data[0][:-1])
        inch = float(sum(Fraction(s) for s in split_data[1][:-1].split()))
        mm = (ft * 12 + inch) * 25.4
        return mm
    elif '"' in data:
        inch = float(sum(Fraction(s) for s in data[:-1].split()))
        mm = inch * 25.4
        return mm
    elif '\'' in data:
        ft = float(data[:-1])
        mm = (ft * 12) * 25.4
        return mm

def populate_spreadsheet():
    print("Populating spreadsheet...")
    csv_files = []
    sep_dwg_names = listdir('drawings')
    for i in range(len(sep_dwg_names)):
        if sep_dwg_names[i][-4:] == '.csv':
            csv_files.append(sep_dwg_names[i])
    # open document
    wb = load_workbook('MTO.xlsx')
    ws = wb.active
    # iterate through csv files
    current_col = 23
    current_row = 6
    done_list = []
    for i in range(len(csv_files)):
        print("File {} of {}".format(i, len(csv_files)))
        # populate data
        with open('drawings/' + csv_files[i], 'r') as file:
            csv_reader = csv.reader(file)
            cleaned_data = []
            # Use this if the table has headers
            try:
                next(csv_reader)
            except StopIteration:
                current_col += 1
                continue
            for line in csv_reader: # The indexes may need to be edited depending on how the drawing BOM is set up. (i.e. i[0] will not always be length in MM, etc.)
                line.pop(0)
                if '"' in line[0] or "'" in line[0]:
                    try:
                        line[0] = read_length(line[0])
                    except ValueError:
                        print("File {} is corrupt".format(csv_files[i]))
                        current_col += 1
                        continue
                try:
                    if line[0]:
                        cleaned_data.append(line)
                    else:
                        continue
                except IndexError:
                    continue
            for i in cleaned_data:
                # if 'MM' in i[0]:
                #     i[0] = i[0][:-3]
                descsize = i[2] + i[1]
                if descsize in done_list:
                    for j in range(5, current_row):
                        if ws.cell(row=j, column=1).value == i[2] and ws.cell(row=j, column=2).value == i[1]:
                            try:
                                ws.cell(row=j, column=current_col, value=float(i[0]))
                            except ValueError:
                                ws.cell(row=j, column=current_col, value=i[0])
                            break
                else:
                    done_list.append(descsize)
                    if i[2].startswith('FLANGE'):
                        ws.cell(row=current_row, column=4, value='FLANGE')
                    elif 'CLAMP' in i[2] or 'TRUNNION' in i[2] or 'BRACE' in i[2]:
                        ws.cell(row=current_row, column=4, value='SUPPORT')
                    elif i[2].startswith('COUPLING'):
                        ws.cell(row=current_row, column=4, value='COUPLING')
                    elif i[2].startswith('GASKET'):
                        ws.cell(row=current_row, column=4, value='GASKET')
                    elif i[2].startswith('BOLT') or i[2].startswith('WASHER'):
                        ws.cell(row=current_row, column=4, value='HARDWARE')
                    elif i[2].startswith('TEE') or i[2].startswith('LAT') or i[2].startswith('REDUCER') or 'ELBOW' in i[2]:
                        ws.cell(row=current_row, column=4, value='FITTING')
                    elif i[2].startswith('PIPE') or 'NIPPLE' in i[2]:
                        ws.cell(row=current_row, column=4, value='PIPE')
                    elif 'VALVE' in i[2]:
                        ws.cell(row=current_row, column=4, value='VALVE')
                    elif 'BEND' in i[2]:
                        ws.cell(row=current_row, column=4, value='BEND')

                    ws.cell(row=current_row, column=1, value=i[2])
                    ws.cell(row=current_row, column=2, value=i[1])
                    try:
                        ws.cell(row=current_row, column=current_col, value=float(i[0]))
                    except ValueError:
                        ws.cell(row=current_row, column=current_col, value=i[0])
                    current_row += 1

        current_col += 1
    
    wb.save('MTO.xlsx')

    print("Done.")

create_spreadhseet()
# convert_to_csv()
populate_spreadsheet()
print("All finished!")