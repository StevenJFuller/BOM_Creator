from os import listdir
import tabula
import csv
from PyPDF2 import PdfFileReader, PdfFileWriter
from openpyxl import Workbook, load_workbook

def seperate_drawings(): # seperates drawings so that all drawings are single page only. Adds sheet number.
    drawing_names = listdir('drawings')
    for i in range(len(drawing_names)):
        inputpdf = PdfFileReader(open('drawings/' + drawing_names[i], 'rb'))
        for j in range(inputpdf.numPages):
            output = PdfFileWriter()
            output.addPage(inputpdf.getPage(j))
            with open('drawings_seperated/' + drawing_names[i][:-4] + "_sheet_" + str(j + 1) + '.pdf', 'wb') as outputStream:
                output.write(outputStream)

def create_spreadhseet(): # sets up MTO spreadsheet with drawing names (typically ISO numbers) along the top row.
    pdf_files = []
    sep_dwg_names = listdir('drawings_seperated')
    for i in range(len(sep_dwg_names)):
        if sep_dwg_names[i][-4:] == '.pdf':
            pdf_files.append(sep_dwg_names[i][:-4])
    wb = load_workbook('Blank MTO r1.xlsx')
    ws = wb.active
    current_column = 23
    for i in range(len(pdf_files)):
        ws.cell(row=5, column=current_column, value=pdf_files[i])
        current_column += 1

    wb.save('MTO.xlsx')

def convert_to_csv(): # converts seperated drawings' BOM's to csv format --note: 'area' may need to be manipulated using tabula-java to define BOM list on drawing
    tabula.convert_into_by_batch("drawings_seperated", output_format="csv", lattice=True, area=(57.88,1154.442,788.22,1674.309))

def populate_spreadsheet():
    csv_files = []
    sep_dwg_names = listdir('drawings_seperated')
    for i in range(len(sep_dwg_names)):
        if sep_dwg_names[i][-4:] == '.csv':
            csv_files.append(sep_dwg_names[i])
    # open document
    wb = load_workbook('MTO.xlsx')
    ws = wb.active
    # iterate through csv files
    current_col = 23
    current_row = 6
    for i in range(len(csv_files)):
        # populate data
        with open('drawings_seperated/' + csv_files[i], 'r') as file:
            csv_reader = csv.reader(file)
            cleaned_data = []
            next(csv_reader)
            for i in csv_reader:
                i.pop(0)
                if i[0]:
                    cleaned_data.append(i)

            for i in cleaned_data:
                if 'MM' in i[0]:
                    i[0] = i[0][:-3]
                for j in range(5, current_row):
                    if ws.cell(row=j, column=1).value == i[2] and ws.cell(row=j, column=2).value == i[1]:
                        ws.cell(row=j, column=current_col, value=int(i[0]))
                        break
                # Error: This is making duplicates. Trying to make it so that if description and size are already in the spreadsheet, add QTY to current_col, if not in spreadsheet, add them
                if i[2].startswith('FLANGE'):
                    ws.cell(row=current_row, column=4, value='FLANGE')
                elif i[2].startswith('COUPLING'):
                    ws.cell(row=current_row, column=4, value='COUPLING')
                elif i[2].startswith('GASKET'):
                    ws.cell(row=current_row, column=4, value='GASKET')
                elif i[2].startswith('BOLT') or i[2].startswith('WASHER'):
                    ws.cell(row=current_row, column=4, value='HARDWARE')
                elif i[2].startswith('TEE') or i[2].startswith('LAT') or i[2].startswith('REDUCER'):
                    ws.cell(row=current_row, column=4, value='FITTING')
                elif i[2].startswith('PIPE'):
                    ws.cell(row=current_row, column=4, value='PIPE')
                elif 'VALVE' in i[2]:
                    ws.cell(row=current_row, column=4, value='VALVE')
                elif 'BEND' in i[2]:
                    ws.cell(row=current_row, column=4, value='BEND')

                ws.cell(row=current_row, column=1, value=i[2])
                ws.cell(row=current_row, column=2, value=i[1])
                ws.cell(row=current_row, column=current_col, value=int(i[0]))
                current_row += 1

        current_col += 1
    
    wb.save('MTO.xlsx')

seperate_drawings()
create_spreadhseet()
convert_to_csv()
populate_spreadsheet()